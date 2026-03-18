import whisper
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import subprocess
import os
os.chdir(r"C:\Users\Sangam Pal\OneDrive\Desktop\AutoEIT-Test")

# ── Config ──────────────────────────────────────────────────
MODEL_SIZE = "medium"   # ya "large-v3" agar GPU/time ho

PARTICIPANTS = {
    "038010": {"file": "data/038010_EIT-2A.mp3", "eit_start": 148},
    "038011": {"file": "data/038011_EIT-1A.mp3", "eit_start": 148},
    "038012": {"file": "data/038012_EIT-2A.mp3", "eit_start": 720},
    "038015": {"file": "data/038015_EIT-1A.mp3", "eit_start": 148},
}

SHEET_MAP = {
    "038010": "38010-2A",
    "038011": "38011-1A",
    "038012": "38012-2A",
    "038015": "38015-1A",
}

STIMULI = [
    (1,  "Quiero cortarme el pelo",                                7),
    (2,  "El libro está en la mesa",                               7),
    (3,  "El carro lo tiene Pedro",                                8),
    (4,  "Él se ducha cada mañana",                                9),
    (5,  "¿Qué dice usted que va a hacer hoy?",                    9),
    (6,  "Dudo que sepa manejar muy bien",                        10),
    (7,  "Las calles de esta ciudad son muy anchas",              11),
    (8,  "Puede que llueva mañana todo el día",                   12),
    (9,  "Las casas son muy bonitas pero caras",                  12),
    (10, "Me gustan las películas que acaban bien",               12),
    (11, "El chico con el que yo salgo es español",               13),
    (12, "Después de cenar me fui a dormir tranquilo",            13),
    (13, "Quiero una casa en la que vivan mis animales",          14),
    (14, "A nosotros nos fascinan las fiestas grandiosas",        14),
    (15, "Ella sólo bebe cerveza y no come nada",                 15),
    (16, "Me gustaría que el precio de las casas bajara",         15),
    (17, "Cruza a la derecha y después sigue todo recto",         15),
    (18, "Ella ha terminado de pintar su apartamento",            15),
    (19, "Me gustaría que empezara a hacer más calor pronto",     15),
    (20, "El niño al que se le murió el gato está triste",        16),
    (21, "Una amiga mía cuida a los niños de mi vecino",          16),
    (22, "El gato que era negro fue perseguido por el perro",     16),
    (23, "Antes de poder salir él tiene que limpiar su cuarto",   16),
    (24, "La cantidad de personas que fuman ha disminuido",       17),
    (25, "Después de llegar a casa del trabajo tomé la cena",     17),
    (26, "El ladrón al que atrapó la policía era famoso",         17),
    (27, "Le pedí a un amigo que me ayudara con la tarea",        17),
    (28, "El examen no fue tan difícil como me habían dicho",     17),
    (29, "¿Serías tan amable de darme el libro que está en la mesa?", 18),
    (30, "Hay mucha gente que no toma nada para el desayuno",     17),
]

# ── Step 1: MP3 → WAV (16kHz mono) ─────────────────────────
def convert_to_wav(mp3_path, wav_path):
    subprocess.run([
        "ffmpeg", "-i", mp3_path,
        "-ar", "16000", "-ac", "1",
        wav_path, "-y", "-loglevel", "quiet"
    ], check=True)
    print(f"  Converted: {mp3_path} → {wav_path}")

# ── Step 2: Segment audio into 30 response windows ──────────
def get_response_windows(wav_path, eit_start_sec):
    """
    Use ffmpeg silence detection to find 30 participant response windows.
    Returns list of (start_sec, end_sec) tuples.
    """
    import re

    result = subprocess.run([
        "ffmpeg", "-i", wav_path,
        "-af", "silencedetect=noise=-35dB:d=1.5",
        "-f", "null", "-"
    ], capture_output=True, text=True)

    output = result.stderr
    silence_starts = [float(x) for x in re.findall(r"silence_start: ([\d.]+)", output)]
    silence_ends   = [float(x) for x in re.findall(r"silence_end: ([\d.]+)",   output)]

    # Build silence intervals after EIT onset
    silences = [(s, e) for s, e in zip(silence_starts, silence_ends)
                if s > eit_start_sec - 5]

    # Speech = gaps between silences
    windows = []
    for i in range(len(silences) - 1):
        speech_start = silences[i][1]
        speech_end   = silences[i + 1][0]
        duration     = speech_end - speech_start
        if duration >= 0.5:
            windows.append((speech_start, speech_end))

    return windows[:30]   # Take first 30

# ── Step 3: Extract each window as a WAV file ────────────────
def extract_segment(wav_path, start, end, out_path):
    subprocess.run([
        "ffmpeg", "-i", wav_path,
        "-ss", str(max(0, start - 0.3)),
        "-to", str(end + 0.3),
        out_path, "-y", "-loglevel", "quiet"
    ], check=True)

# ── Step 4: Transcribe with Whisper ─────────────────────────
def transcribe_segment(model, wav_path):
    result = model.transcribe(
        wav_path,
        language="es",          # Force Spanish
        task="transcribe",
        beam_size=5,
        temperature=0,
        no_speech_threshold=0.6,
    )
    text = result["text"].strip()
    # If Whisper detected no speech
    if not text or text.lower() in [".", ",", "..."]:
        return "[silence]"
    return text

# ── Step 5: Write to Excel ───────────────────────────────────
def write_excel(all_results, template_path, output_path):
    wb = openpyxl.load_workbook(template_path)

    hdr  = Font(name="Arial", bold=True,  size=10)
    body = Font(name="Arial",             size=10)
    grey = Font(name="Arial",             size=10, color="888888", italic=True)

    blue_fill = PatternFill("solid", start_color="BDD7EE")
    alt_fill  = PatternFill("solid", start_color="F2F2F2")
    sil_fill  = PatternFill("solid", start_color="EFEFEF")

    center = Alignment(horizontal="center", vertical="center")
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    for pid, items in all_results.items():
        ws = wb[SHEET_MAP[pid]]

        # Headers
        ws["A1"] = "Sentence";   ws["A1"].font = hdr; ws["A1"].fill = blue_fill; ws["A1"].alignment = center
        ws["B1"] = "Stimulus";   ws["B1"].font = hdr; ws["B1"].fill = blue_fill; ws["B1"].alignment = center
        ws["C1"] = "Transcription"; ws["C1"].font = hdr; ws["C1"].fill = blue_fill; ws["C1"].alignment = center
        ws["D1"] = "Timing (s)"; ws["D1"].font = hdr; ws["D1"].fill = blue_fill; ws["D1"].alignment = center

        ws.column_dimensions["A"].width = 9
        ws.column_dimensions["B"].width = 50
        ws.column_dimensions["C"].width = 52
        ws.column_dimensions["D"].width = 18
        ws.freeze_panes = "A2"

        for item_num, stimulus, syllables, transcription, start_s, end_s in items:
            row = item_num + 1
            dur = round(end_s - start_s, 1) if start_s and end_s else 0
            timing = f"{start_s:.1f}–{end_s:.1f} ({dur}s)" if start_s else "—"
            is_sil = transcription == "[silence]"

            ws.cell(row, 1).value = item_num
            ws.cell(row, 2).value = f"{stimulus} ({syllables})"
            ws.cell(row, 3).value = transcription
            ws.cell(row, 4).value = timing

            no_fill = PatternFill(fill_type=None)
            for col in range(1, 5):
                c = ws.cell(row, col)
                c.font      = grey if is_sil else body
                c.fill      = sil_fill if is_sil else (alt_fill if row % 2 == 0 else no_fill)
                c.alignment = center if col in (1, 4) else left

        print(f"  Excel: {pid} → sheet '{SHEET_MAP[pid]}' written")

    wb.save(output_path)
    print(f"\n✓ Saved: {output_path}")

# ── MAIN ─────────────────────────────────────────────────────
def main():
    print("Loading Whisper model:", MODEL_SIZE)
    model = whisper.load_model(MODEL_SIZE)
    print("Model loaded.\n")

    all_results = {}

    for pid, config in PARTICIPANTS.items():
        print(f"\n{'='*55}")
        print(f"Participant: {pid}")
        print(f"{'='*55}")

        mp3_path = config["file"]
        eit_start = config["eit_start"]

        # Convert MP3 to WAV
        wav_path = f"outputs/{pid}_full.wav"
        convert_to_wav(mp3_path, wav_path)

        # Get 30 response windows
        print("  Detecting speech windows...")
        windows = get_response_windows(wav_path, eit_start)
        print(f"  Found {len(windows)} windows")

        # Transcribe each window
        results = []
        for i, (start, end) in enumerate(windows):
            item_num = i + 1
            seg_path = f"outputs/{pid}_seg_{item_num:02d}.wav"

            extract_segment(wav_path, start, end, seg_path)
            text = transcribe_segment(model, seg_path)

            stimulus, syllables = STIMULI[i][1], STIMULI[i][2]
            results.append((item_num, stimulus, syllables, text, start, end))
            print(f"  Item {item_num:02d}: {text[:60]}")

        # Pad missing items with silence
        while len(results) < 30:
            i = len(results)
            results.append((i+1, STIMULI[i][1], STIMULI[i][2], "[silence]", None, None))

        all_results[pid] = results

    # Write Excel
    print(f"\n{'='*55}")
    print("Writing Excel output...")
    write_excel(
        all_results,
        template_path="data/AutoEIT Sample Audio for Transcribing.xlsx",
        output_path="outputs/AutoEIT_Transcriptions_Completed.xlsx"
    )

if __name__ == "__main__":
    main()